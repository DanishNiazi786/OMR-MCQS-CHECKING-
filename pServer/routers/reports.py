from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from models.report import ReportCreate, ReportResponse
from datetime import datetime

router = APIRouter()

def get_database():
    from main import app
    return app.state.database

@router.post("/excel/{exam_id}")
async def generate_excel_report(exam_id: str, db=Depends(get_database)):
    try:
        # Verify exam exists
        exam = await db.exams.find_one({"examId": exam_id})
        if not exam:
            raise HTTPException(status_code=404, detail="Exam not found")

        # Get responses
        cursor = db.responses.find({"examId": exam_id}).sort("studentId", 1)
        responses = await cursor.to_list(length=None)

        # Create workbook
        wb = Workbook()
        
        # Individual Results Sheet
        ws_results = wb.active
        ws_results.title = "Individual Results"
        
        # Headers
        headers = ['Student ID', 'Score', 'Accuracy (%)', 'Correct', 'Incorrect', 'Blank', 'Multiple Marks', 'Processed At']
        ws_results.append(headers)
        
        # Style headers
        for cell in ws_results[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # Add data rows
        for response in responses:
            ws_results.append([
                response["studentId"],
                response["score"],
                round(response["accuracy"], 2),
                response["correctAnswers"],
                response["incorrectAnswers"],
                response["blankAnswers"],
                response["multipleMarks"],
                response["processedAt"].strftime("%Y-%m-%d")
            ])

        # Auto-fit columns
        for column in ws_results.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_results.column_dimensions[column_letter].width = adjusted_width

        # Statistics Sheet
        ws_stats = wb.create_sheet("Statistics")
        stats = await calculate_exam_stats(exam_id, db)
        
        ws_stats.append(["Exam Statistics"])
        ws_stats.append(["Total Students", stats["totalStudents"]])
        ws_stats.append(["Average Score", stats["averageScore"]])
        ws_stats.append(["Highest Score", stats["highestScore"]])
        ws_stats.append(["Lowest Score", stats["lowestScore"]])
        ws_stats.append(["Passing Rate (%)", stats["passingRate"]])

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Save report record
        report_data = {
            "examId": exam_id,
            "reportType": "Excel",
            "data": stats,
            "generatedAt": datetime.utcnow(),
            "generatedBy": "System"
        }
        await db.reports.insert_one(report_data)

        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={exam['name']}_Report.xlsx"}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to generate Excel report")

@router.post("/pdf/{exam_id}")
async def generate_pdf_report(exam_id: str, db=Depends(get_database)):
    try:
        # Verify exam exists
        exam = await db.exams.find_one({"examId": exam_id})
        if not exam:
            raise HTTPException(status_code=404, detail="Exam not found")

        # Get responses and stats
        cursor = db.responses.find({"examId": exam_id}).sort("studentId", 1)
        responses = await cursor.to_list(length=None)
        stats = await calculate_exam_stats(exam_id, db)

        # Create PDF
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Title
        p.setFont("Helvetica-Bold", 20)
        p.drawCentredText(width/2, height-50, "EFSoft OMR Report")
        p.setFont("Helvetica", 14)
        p.drawCentredText(width/2, height-80, f"Exam: {exam['name']}")
        p.drawCentredText(width/2, height-100, f"Generated: {datetime.now().strftime('%Y-%m-%d')}")

        # Statistics Section
        y_pos = height - 150
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, y_pos, "Exam Statistics")
        p.setFont("Helvetica", 12)
        y_pos -= 25
        p.drawString(50, y_pos, f"Total Students: {stats['totalStudents']}")
        y_pos -= 20
        p.drawString(50, y_pos, f"Average Score: {stats['averageScore']}")
        y_pos -= 20
        p.drawString(50, y_pos, f"Highest Score: {stats['highestScore']}")
        y_pos -= 20
        p.drawString(50, y_pos, f"Lowest Score: {stats['lowestScore']}")
        y_pos -= 20
        p.drawString(50, y_pos, f"Passing Rate: {stats['passingRate']}%")

        # Score Distribution
        y_pos -= 40
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y_pos, "Score Distribution")
        p.setFont("Helvetica", 10)
        y_pos -= 20
        for dist in stats["scoreDistribution"]:
            p.drawString(50, y_pos, f"{dist['range']}: {dist['count']} students ({round(dist['percentage'])}%)")
            y_pos -= 15

        # Individual Results Table
        p.showPage()
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height-50, "Individual Results")
        p.setFont("Helvetica", 8)

        # Table headers
        table_top = height - 100
        p.drawString(50, table_top, "Student ID")
        p.drawString(150, table_top, "Score")
        p.drawString(200, table_top, "Accuracy")
        p.drawString(250, table_top, "Correct")
        p.drawString(300, table_top, "Incorrect")
        p.drawString(350, table_top, "Blank")
        p.drawString(400, table_top, "Multiple")

        # Table rows
        current_y = table_top - 20
        for response in responses[:40]:  # Limit to 40 results per page
            if current_y < 100:  # Start new page if needed
                p.showPage()
                current_y = height - 50

            p.drawString(50, current_y, str(response["studentId"]))
            p.drawString(150, current_y, str(response["score"]))
            p.drawString(200, current_y, f"{round(response['accuracy'])}%")
            p.drawString(250, current_y, str(response["correctAnswers"]))
            p.drawString(300, current_y, str(response["incorrectAnswers"]))
            p.drawString(350, current_y, str(response["blankAnswers"]))
            p.drawString(400, current_y, str(response["multipleMarks"]))

            current_y -= 15

        p.save()
        buffer.seek(0)

        # Save report record
        report_data = {
            "examId": exam_id,
            "reportType": "PDF",
            "data": stats,
            "generatedAt": datetime.utcnow(),
            "generatedBy": "System"
        }
        await db.reports.insert_one(report_data)

        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={exam['name']}_Report.pdf"}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to generate PDF report")

@router.get("/history/{exam_id}")
async def get_report_history(exam_id: str, db=Depends(get_database)):
    try:
        cursor = db.reports.find({"examId": exam_id}).sort("generatedAt", -1)
        reports = await cursor.to_list(length=None)
        
        # Convert ObjectId to string
        for report in reports:
            report['_id'] = str(report['_id'])
        
        return reports
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to fetch report history")

async def calculate_exam_stats(exam_id: str, db):
    cursor = db.responses.find({"examId": exam_id})
    responses = await cursor.to_list(length=None)
    
    exam = await db.exams.find_one({"examId": exam_id})

    if not responses:
        return {
            "totalStudents": 0,
            "averageScore": 0,
            "highestScore": 0,
            "lowestScore": 0,
            "passingRate": 0,
            "scoreDistribution": [],
            "questionAnalysis": []
        }

    scores = [r["score"] for r in responses]
    total_students = len(responses)
    average_score = sum(scores) / total_students
    highest_score = max(scores)
    lowest_score = min(scores)
    
    passing_score = exam.get("settings", {}).get("passingScore", 60)
    passing_count = len([s for s in scores if (s / exam["numQuestions"]) * 100 >= passing_score])
    passing_rate = (passing_count / total_students) * 100

    # Score distribution
    ranges = [
        {"min": 0, "max": 20, "label": "0-20%"},
        {"min": 21, "max": 40, "label": "21-40%"},
        {"min": 41, "max": 60, "label": "41-60%"},
        {"min": 61, "max": 80, "label": "61-80%"},
        {"min": 81, "max": 100, "label": "81-100%"}
    ]

    score_distribution = []
    for range_item in ranges:
        count = len([s for s in scores if range_item["min"] <= (s / exam["numQuestions"]) * 100 <= range_item["max"]])
        score_distribution.append({
            "range": range_item["label"],
            "count": count,
            "percentage": (count / total_students) * 100
        })

    return {
        "totalStudents": total_students,
        "averageScore": round(average_score, 2),
        "highestScore": highest_score,
        "lowestScore": lowest_score,
        "passingRate": round(passing_rate, 2),
        "scoreDistribution": score_distribution,
        "questionAnalysis": []  # Simplified for now
    }